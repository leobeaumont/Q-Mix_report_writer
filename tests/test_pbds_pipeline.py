"""
Tests for the PBDS chunk -> node-name -> k-hop neighbourhood pipeline
(qmix_report_writer/tools/pbds).

The real PBDSManager builds its dependency graph from the LEANDREA workbook
(formulas + value columns). That workbook is not available to tests, so these
tests inject a small hand-built fixture graph + descriptions straight into a real
PBDSManager (the loader is skipped when `_graph` is already set). Everything else
is the production code path: BM25 shortlisting, the LLM verifier's parsing/
multi-pick logic, real networkx k-hop traversal, and expand()'s priority ranking,
dedup and truncation.

The deterministic tests mock the LLM (no Ollama, no workbook needed). One live
test drives the configured Ollama model against the real parameter descriptions
in `parameter names and descriptions.xlsx`; it SKIPS (does not fail) if Ollama is
unreachable or the sample file is missing.

Run:  .venv/Scripts/python.exe tests/test_pbds_pipeline.py
"""

import asyncio
import os
import sys
import urllib.request

sys.path.insert(0, ".")

from unittest.mock import AsyncMock, MagicMock

import networkx as nx

from qmix_report_writer.tools.pbds import Candidate, NodeMatcher, PBDSManager

SAMPLE_XLSX = "parameter names and descriptions.xlsx"


# ---------------------------------------------------------------------------
# Fixture: a small dependency graph injected into a real PBDSManager.
#
#   Core_inlet_temperature_F28 -> Core_power_F26 -> Max_fuel_temperature_F10
#   Pellet_diameter_F8 ----------------------------> Max_fuel_temperature_F10
#   Max_fuel_temperature_F10 -> Cladding_thickness_F15
#   Vessel_size_F36 -> Reactor_size_F37            (disjoint sub-graph)
#
# Edge src -> dst means dst is computed from src (formula stored on the edge).
# ---------------------------------------------------------------------------

_DESCRIPTIONS = {
    "Pellet_diameter_F8": "Pellet diameter",
    "Max_fuel_temperature_F10": "Max fuel temperature",
    "Cladding_thickness_F15": "Cladding thickness",
    "Core_power_F26": "Core power",
    "Core_inlet_temperature_F28": "Core inlet temperature",
    "Maximum_core_power_F44": "Maximum core power",
    "Vessel_size_F36": "Vessel size",
    "Reactor_size_F37": "Reactor size",
}

_EDGES = [
    ("Pellet_diameter_F8", "Max_fuel_temperature_F10", "=F(C8)"),
    ("Core_power_F26", "Max_fuel_temperature_F10", "=G(C26)"),
    ("Core_inlet_temperature_F28", "Core_power_F26", "=H(C28)"),
    ("Max_fuel_temperature_F10", "Cladding_thickness_F15", "=I(C10)"),
    ("Vessel_size_F36", "Reactor_size_F37", "=J(C36)"),
]


def _build_manager(default_k: int = 2) -> PBDSManager:
    """A real PBDSManager with the fixture graph/descriptions injected (no workbook)."""
    g = nx.DiGraph()
    for name in _DESCRIPTIONS:
        row = int(name.rsplit("_F", 1)[1])
        g.add_node(name, owner="MGR", row=row, row_type="calculated_parameter")
    for src, dst, formula in _EDGES:
        g.add_edge(src, dst, formulas=[formula])

    m = PBDSManager("<in-memory>", default_k=default_k)
    m._graph = g                      # skips the workbook loader
    m._descriptions = dict(_DESCRIPTIONS)
    return m


def _mock_llm(node_ids_json: str):
    """A MagicMock LLM whose agen() returns a fixed JSON string."""
    llm = MagicMock()
    llm.agen = AsyncMock(return_value=node_ids_json)
    return llm


# ---------------------------------------------------------------------------
# 1. Graph traversal: sources/effects, hops, edge formulas.
# ---------------------------------------------------------------------------

async def test_manager_graph_traversal():
    m = _build_manager()
    nb = m.neighborhood("Max_fuel_temperature_F10", k=2)

    sources = {c.node: c for c in nb["sources"]}
    effects = {c.node: c for c in nb["effects"]}

    assert set(sources) == {
        "Pellet_diameter_F8",
        "Core_power_F26",
        "Core_inlet_temperature_F28",
    }, sources
    assert set(effects) == {"Cladding_thickness_F15"}, effects

    # Hop distances: direct predecessors at 1, the two-step one at 2.
    assert sources["Pellet_diameter_F8"].hops == 1
    assert sources["Core_power_F26"].hops == 1
    assert sources["Core_inlet_temperature_F28"].hops == 2

    # Edge formulas travel with the connection.
    assert sources["Pellet_diameter_F8"].formulas == ["=F(C8)"]
    assert effects["Cladding_thickness_F15"].formulas == ["=I(C10)"]

    # k=1 must NOT reach the two-hop source.
    nb1 = m.neighborhood("Max_fuel_temperature_F10", k=1)
    assert {c.node for c in nb1["sources"]} == {"Pellet_diameter_F8", "Core_power_F26"}
    print("PASS  test_manager_graph_traversal")


# ---------------------------------------------------------------------------
# 2. Error handling on the traversal API.
# ---------------------------------------------------------------------------

async def test_traversal_error_handling():
    m = _build_manager()

    for bad in [
        lambda: m.connected_nodes("DOES_NOT_EXIST"),
        lambda: m.connected_nodes("Core_power_F26", k=0),
        lambda: m.connected_nodes("Core_power_F26", direction="upstream"),
    ]:
        try:
            bad()
            assert False, "expected an exception"
        except (KeyError, ValueError):
            pass
    print("PASS  test_traversal_error_handling")


# ---------------------------------------------------------------------------
# 3. Layer 1: BM25 shortlist (on-topic hit, off-topic empties, no LLM).
# ---------------------------------------------------------------------------

async def test_shortlist_bm25():
    nm = NodeMatcher(_build_manager())

    on = nm.shortlist("We must keep the max fuel temperature within limits.")
    assert any(c.node == "Max_fuel_temperature_F10" for c in on), on

    # No description token appears -> nothing scores -> no LLM call wasted.
    off = nm.shortlist("The cafeteria parking lot was repaved on Sunday.")
    assert off == [], off
    print("PASS  test_shortlist_bm25")


# ---------------------------------------------------------------------------
# 4. Layer 2: verifier multi-pick + lenient (fenced) JSON + score-order.
# ---------------------------------------------------------------------------

async def test_shortlist_dedupes_by_description():
    # Two nodes share "Reactor core temperature" (as the real workbook reuses
    # labels). For the query below they tie for the top score, so without dedup
    # they would take both of the top_k=2 slots and squeeze out the distinct
    # "Reactor core pressure". Dedup must keep one of them and free the slot.
    descriptions = {
        "Reactor_core_temperature_A1": "Reactor core temperature",
        "Reactor_core_temperature_A2": "Reactor core temperature",  # duplicate label
        "Reactor_core_pressure_B1": "Reactor core pressure",
        "Alpha_C1": "Alpha metric one",   # fillers for sane IDF
        "Beta_C2": "Beta metric two",
    }

    class StubManager:
        def descriptions(self):
            return descriptions

    nm = NodeMatcher(StubManager(), top_k=2, min_score_ratio=0.0)
    sl = nm.shortlist("reactor core temperature")
    labels = [c.description for c in sl]
    assert labels.count("Reactor core temperature") == 1, sl        # deduped
    assert "Reactor core pressure" in labels, sl                    # slot freed for it
    print("PASS  test_shortlist_dedupes_by_description")


async def test_verify_multipick_and_lenient_json():
    nm = NodeMatcher(_build_manager())
    cands = [
        Candidate("Max_fuel_temperature_F10", "Max fuel temperature", 10.0),
        Candidate("Core_power_F26", "Core power", 5.0),
    ]

    # Model wraps a JSON array in a markdown fence and lists ids out of order.
    fenced = (
        "```json\n"
        '{"node_ids": ["Core_power_F26", "Max_fuel_temperature_F10"], "reason": "both"}\n'
        "```"
    )
    out = await nm.verify("chunk", cands, _mock_llm(fenced))
    # Both confirmed, returned in shortlist (score) order, fence tolerated.
    assert [c.node for c in out] == ["Max_fuel_temperature_F10", "Core_power_F26"], out

    # Empty selection -> empty result.
    none = await nm.verify("chunk", cands, _mock_llm('{"node_ids": [], "reason": "none"}'))
    assert none == [], none

    # Garbage that yields no JSON object -> empty result (no crash).
    junk = await nm.verify("chunk", cands, _mock_llm("the model said nothing useful"))
    assert junk == [], junk
    print("PASS  test_verify_multipick_and_lenient_json")


# ---------------------------------------------------------------------------
# 5. expand(): single match -> neighbours, then truncation by priority.
# ---------------------------------------------------------------------------

async def test_expand_single_match_and_truncation():
    m = _build_manager()
    llm = _mock_llm('{"node_ids": ["Max_fuel_temperature_F10"], "reason": "x"}')

    res = await NodeMatcher(m, max_neighbors=None).expand("max fuel temperature", llm, k=2)
    assert [mn["node"] for mn in res["matched_nodes"]] == ["Max_fuel_temperature_F10"]

    got = {(n.node, n.direction) for n in res["neighbors"]}
    assert got == {
        ("Pellet_diameter_F8", "sources"),
        ("Core_power_F26", "sources"),
        ("Core_inlet_temperature_F28", "sources"),
        ("Cladding_thickness_F15", "effects"),
    }, got

    # Same origin (equal match_score) -> rank by (hops, node); cap keeps the top 2.
    res2 = await NodeMatcher(m, max_neighbors=2).expand("max fuel temperature", llm, k=2)
    assert [n.node for n in res2["neighbors"]] == [
        "Cladding_thickness_F15",
        "Core_power_F26",
    ], [n.node for n in res2["neighbors"]]
    print("PASS  test_expand_single_match_and_truncation")


# ---------------------------------------------------------------------------
# 6. expand(): multiple matches -> better-match neighbours win priority.
# ---------------------------------------------------------------------------

async def test_expand_multi_match_priority():
    m = _build_manager()
    llm = _mock_llm(
        '{"node_ids": ["Max_fuel_temperature_F10", "Vessel_size_F36"], "reason": "x"}'
    )
    res = await NodeMatcher(m, max_neighbors=None).expand(
        "max fuel temperature and vessel size", llm, k=2
    )

    scores = {mn["node"]: mn["score"] for mn in res["matched_nodes"]}
    assert set(scores) == {"Max_fuel_temperature_F10", "Vessel_size_F36"}, scores
    assert scores["Max_fuel_temperature_F10"] != scores["Vessel_size_F36"], scores

    neighbors = res["neighbors"]
    origins = [n.origin for n in neighbors]
    hi = max(scores, key=scores.get)
    lo = min(scores, key=scores.get)

    # Every neighbour of the higher-scoring match precedes every neighbour of the lower one.
    last_hi = max(i for i, o in enumerate(origins) if o == hi)
    first_lo = min(i for i, o in enumerate(origins) if o == lo)
    assert last_hi < first_lo, origins

    # Output respects the documented ordering key and has no duplicate (node, direction).
    keys = [(-n.match_score, n.hops) for n in neighbors]
    assert keys == sorted(keys), keys
    pairs = [(n.node, n.direction) for n in neighbors]
    assert len(pairs) == len(set(pairs)), pairs

    # Matched nodes are never returned as neighbours of one another.
    assert not ({"Max_fuel_temperature_F10", "Vessel_size_F36"} & {n.node for n in neighbors})
    print("PASS  test_expand_multi_match_priority")


# ---------------------------------------------------------------------------
# 7. expand(): no confirmed node -> empty neighbours, no traversal.
# ---------------------------------------------------------------------------

async def test_expand_no_match():
    m = _build_manager()
    # Off-topic chunk never reaches the LLM (empty shortlist).
    res = await NodeMatcher(m).expand("the cafeteria parking lot", _mock_llm("{}"), k=2)
    assert res["matched_nodes"] == [], res
    assert res["neighbors"] == [], res
    print("PASS  test_expand_no_match")


# ---------------------------------------------------------------------------
# 8. LIVE: real Ollama verifier against the real parameter descriptions.
# ---------------------------------------------------------------------------

def _ollama_up() -> bool:
    try:
        urllib.request.urlopen("http://localhost:11434/api/tags", timeout=5)
        return True
    except Exception:
        return False


def _load_sample_descriptions():
    """name -> description from the sample workbook, or None if it is missing."""
    if not os.path.exists(SAMPLE_XLSX):
        return None
    import openpyxl

    wb = openpyxl.load_workbook(SAMPLE_XLSX, data_only=True)
    ws = wb["LEANDREA Parameters"]
    descs = {}
    for r in range(5, ws.max_row + 1):
        name, desc = ws.cell(r, 2).value, ws.cell(r, 3).value
        if name and desc:
            descs[str(name).strip()] = str(desc).strip()
    return descs


async def test_live_amatch():
    if not _ollama_up():
        print("SKIP  test_live_amatch: Ollama not reachable on :11434")
        return "skip"
    descs = _load_sample_descriptions()
    if not descs:
        print(f"SKIP  test_live_amatch: {SAMPLE_XLSX!r} not found")
        return "skip"

    from qmix_report_writer.llm.ollama_chat import OllamaChat
    from qmix_report_writer.utils.config import get_llm_config

    m = PBDSManager("<live>")
    m._graph = nx.DiGraph()          # amatch never touches the graph
    m._descriptions = descs
    nm = NodeMatcher(m, top_k=6)
    llm = OllamaChat(model_name=get_llm_config().get("default_model"), temperature=0.2)

    on = await nm.amatch(
        "The fuel pellet diameter and the MOX fuel plutonium enrichment together "
        "drive the maximum fuel temperature reached during irradiation.",
        llm,
    )
    assert len(on) >= 1, f"live verifier confirmed nothing on an on-topic chunk: {on}"

    off = await nm.amatch(
        "The turbine hall sits outside the containment building near the river.", llm
    )
    assert off == [], f"live verifier wrongly matched an off-topic chunk: {[c.node for c in off]}"

    print(f"PASS  test_live_amatch (on-topic matched {[c.node for c in on]})")
    return "pass"


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------

async def _run_all():
    tests = [
        test_manager_graph_traversal,
        test_traversal_error_handling,
        test_shortlist_bm25,
        test_shortlist_dedupes_by_description,
        test_verify_multipick_and_lenient_json,
        test_expand_single_match_and_truncation,
        test_expand_multi_match_priority,
        test_expand_no_match,
        test_live_amatch,
    ]
    passed = failed = skipped = 0
    for test in tests:
        try:
            result = await test()
            if result == "skip":
                skipped += 1
            else:
                passed += 1
        except Exception as exc:
            print(f"FAIL  {test.__name__}: {exc}")
            import traceback
            traceback.print_exc()
            failed += 1

    print(f"\n{passed} passed, {failed} failed, {skipped} skipped.")
    return failed


if __name__ == "__main__":
    sys.exit(asyncio.run(_run_all()))
