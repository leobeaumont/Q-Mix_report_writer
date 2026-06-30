"""
Trimmed from the PBDS repository (original file: pbds_pareto_demo_core_v8.py).

Only the dependency-graph path is kept: read the main parameter sheet, classify
its rows, and build a directed dependency graph whose edges carry the connecting
formula(s). The Pareto exploration / LHS sampling / workbook-export machinery from
the original demo has been removed because the report-writer pipeline only needs
the graph for k-hop neighbourhood queries.

This code is not tracking the original repository and requires manual updates.
"""


from __future__ import annotations

__version__ = "pdbs_pareto_demo_core_v8-graph-trim-1.0.0"

import math
import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import networkx as nx
import openpyxl
import pandas as pd

MAIN_SHEET = "LEANDREA Parameters"
HEADER_ROW = 4
DEFAULT_OWNER = "MGR"
DEFAULT_OWNER_DISPLAY = "MGR (default when unassigned)"

_NUMERIC_RE = re.compile(r'^\s*[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[Ee][-+]?\d+)?\s*$')
_CELL_RE = re.compile(r"\$?[A-Z]{1,3}\$?\d+")
_SHEET_CELL_RE = re.compile(r"(?:(?:'([^']+)')|([A-Za-z0-9_ ]+))!\$?[A-Z]{1,3}\$?\d+")
_RANGE_RE = re.compile(r"\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+")
_SHEET_RANGE_RE = re.compile(r"(?:(?:'([^']+)')|([A-Za-z0-9_ ]+))!\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+")
_ROW_REF_RE = re.compile(r'(?<![A-Z0-9_])(?:[A-Z]{1,3})(\d+)(?![A-Z0-9_])')
_SECTION_NAME_HINTS = {
    "constraints", "constraint", "pellets", "pellet", "cladding", "wrapper", "design constraints",
    "fuel bundle", "fuel", "assembly", "assemblies", "coolant", "core", "reactor"
}

def _clean_header(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip().lower().replace("/", " ").replace("_", " ")

def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s.lower() in {"none", "nan", "n/a"}

def _is_formula(v: Any) -> bool:
    return isinstance(v, str) and str(v).startswith("=")

def _is_numericish(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return True
    return bool(_NUMERIC_RE.match(str(v)))

def _to_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        try:
            if math.isnan(v):  # type: ignore[arg-type]
                return None
        except Exception:
            pass
        return float(v)
    s = str(v).strip()
    if s == "" or s.lower() in {"nan", "none", "n/a"}:
        return None
    try:
        return float(s)
    except Exception:
        return None

def _to_display(v: Any) -> Any:
    if _is_blank(v):
        return None
    if isinstance(v, bool):
        return bool(v)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            if math.isnan(v):  # type: ignore[arg-type]
                return None
        except Exception:
            pass
        return v
    return str(v).strip()

def _display_owner(v: Any) -> str:
    if _is_blank(v):
        return DEFAULT_OWNER
    s = str(v).strip()
    if s in {"Unknown", "Unassigned", "Shared"}:
        return DEFAULT_OWNER
    if s.upper().startswith("ALL"):
        return "Shared"
    return s

def _display_owner_ui(owner: Any, defaulted: bool = False) -> str:
    s = _display_owner(owner)
    if defaulted and s == DEFAULT_OWNER:
        return DEFAULT_OWNER_DISPLAY
    return s

def _find_main_columns(ws) -> Dict[str, int]:
    headers = {i: _clean_header(ws.cell(HEADER_ROW, i).value) for i in range(1, ws.max_column + 1)}
    wanted = {}
    for idx, h in headers.items():
        if h == "parameter":
            wanted["parameter"] = idx
        elif h == "low":
            wanted["low"] = idx
        elif h == "high":
            wanted["high"] = idx
        elif "proposed" in h:
            wanted["proposed"] = idx
        elif h == "unit":
            wanted["unit"] = idx
        elif "owner" in h or "lfi" in h:
            wanted["owner"] = idx
        elif h == "description":
            wanted["description"] = idx
        elif h == "constraint formula" or h == "constraint   formula":
            wanted["formula"] = idx
        elif h == "constraint text   formula text" or h == "constraint text formula text":
            wanted["formula_text"] = idx
        elif "constraint" in h or "formula" in h:
            wanted.setdefault("formula", idx)
    required = {"parameter", "low", "high", "proposed"}
    missing = required - set(wanted)
    if missing:
        raise ValueError(f"Main sheet is missing required columns: {sorted(missing)}")
    return wanted

def _cell_in_wide_merged_band(ws, row: int, col: int) -> bool:
    coord = ws.cell(row, col).coordinate
    for rng in ws.merged_cells.ranges:
        if coord in rng and rng.min_row == rng.max_row and (rng.max_col - rng.min_col) >= 4:
            return True
    return False

def _parameter_name_looks_real(s: str) -> bool:
    s = (s or "").strip()
    if s == "" or s.lower() in {"nan", "none"}:
        return False
    if _is_numericish(s):
        return False
    return True

def _is_likely_header_label(s: str) -> bool:
    ss = (s or "").strip().lower()
    if not ss:
        return False
    tokenish = re.sub(r'[_\-]+', ' ', ss)
    words = [w for w in re.split(r'\s+', tokenish) if w]
    if len(words) <= 4 and any(w in _SECTION_NAME_HINTS for w in words):
        return True
    return False

def _humanize_constraint(low_num, high_num, formula_text, formula_raw) -> Optional[str]:
    if low_num is not None and high_num is not None:
        return f"[{low_num:g}, {high_num:g}]"
    if low_num is not None:
        return f"≥ {low_num:g}"
    if high_num is not None:
        return f"≤ {high_num:g}"
    txt = None
    if isinstance(formula_text, str) and formula_text.strip():
        txt = formula_text.strip()
    elif isinstance(formula_raw, str) and formula_raw.strip():
        txt = formula_raw.strip()
    if not txt:
        return None
    if txt == "=FALSE" or txt.upper() == "FALSE":
        return "Workbook logical check"
    return txt

def _synthetic_bounds(low_num: Optional[float], high_num: Optional[float], initial_num: Optional[float], proposed_num: Optional[float], raw_low: Any = None, raw_high: Any = None) -> Tuple[Optional[float], Optional[float], Optional[float], bool]:
    """Return (low, high, initial, synthetic_flag) for Pareto/demo exploration.

    Prefer explicit numeric [low, high]. Otherwise synthesize a conservative local interval around
    a numeric baseline so the demo can still explore a couple of meaningful levers.
    """
    if low_num is not None and high_num is not None and low_num <= high_num:
        init = initial_num if initial_num is not None else proposed_num
        if init is None:
            init = 0.5 * (low_num + high_num)
        init = min(max(init, low_num), high_num)
        return float(low_num), float(high_num), float(init), False

    baseline = None
    if initial_num is not None:
        baseline = initial_num
    elif proposed_num is not None:
        baseline = proposed_num
    elif low_num is not None and high_num is None:
        baseline = low_num
    elif high_num is not None and low_num is None:
        baseline = high_num

    if baseline is None or not math.isfinite(float(baseline)):
        return None, None, None, False

    baseline = float(baseline)
    if baseline == 0:
        delta = 1.0
    else:
        delta = max(abs(baseline) * 0.10, 1e-6)

    if low_num is not None and high_num is None:
        lo = float(low_num)
        hi = max(lo, baseline + delta)
        init = baseline
        return lo, hi, min(max(init, lo), hi), True

    if high_num is not None and low_num is None:
        hi = float(high_num)
        lo = min(hi, baseline - delta)
        init = baseline
        return lo, hi, min(max(init, lo), hi), True

    lo = baseline - delta
    hi = baseline + delta
    return lo, hi, baseline, True

def _safe_cell(ws, row: int, col: Optional[int]):
    if not col:
        return None
    return ws.cell(row, col).value

@dataclass
class WorkbookModel:
    workbook_path: str
    workbook: Any
    memo: Dict[Tuple[str, str], Any]

    @classmethod
    def open(cls, workbook_path: str) -> "WorkbookModel":
        wb = openpyxl.load_workbook(workbook_path, data_only=False)
        return cls(workbook_path, wb, {})

    def close(self):
        try:
            self.workbook.close()
        except Exception:
            pass

    def get_sheet(self, sheet_name: str):
        return self.workbook[sheet_name]

    def get_cell(self, sheet_name: str, addr: str, overrides: Optional[Dict[Tuple[str, str], Any]] = None) -> Any:
        addr = addr.replace("$", "").upper()
        key = (sheet_name, addr)
        ov = overrides or {}
        if key in ov:
            return ov[key]
        if key in self.memo and not ov:
            return self.memo[key]
        ws = self.workbook[sheet_name]
        cell = ws[addr]
        val = cell.value
        if isinstance(val, str) and val.startswith("="):
            out = self.eval_formula(val, sheet_name, overrides=ov)
        else:
            out = self._value_to_scalar(val)
        if not ov:
            self.memo[key] = out
        return out

    def get_range(self, sheet_name: str, start: str, end: str, overrides: Optional[Dict[Tuple[str, str], Any]] = None) -> List[Any]:
        cells = self.workbook[sheet_name][f"{start}:{end}"]
        out = []
        for row in cells:
            for cell in row:
                out.append(self.get_cell(sheet_name, cell.coordinate, overrides=overrides))
        return out

    def _value_to_scalar(self, value: Any) -> Any:
        if isinstance(value, str):
            s = value.strip()
            if s.startswith("="):
                return self.eval_formula(s, MAIN_SHEET)
            num = _to_number(s)
            if num is not None:
                return num
            sl = s.lower()
            if sl == "true":
                return True
            if sl == "false":
                return False
            return s
        if isinstance(value, bool):
            return value
        num = _to_number(value)
        return 0.0 if num is None else num

    def eval_formula(self, formula: str, current_sheet: str, overrides: Optional[Dict[Tuple[str, str], Any]] = None) -> Any:
        expr = formula[1:] if formula.startswith("=") else formula
        expr = expr.replace("^", "**").replace(";", ",")
        placeholders: Dict[str, str] = {}
        ph_counter = 0

        def _ph(replacement: str) -> str:
            nonlocal ph_counter
            key = f"__PH_{ph_counter}__"
            ph_counter += 1
            placeholders[key] = replacement
            return key

        def repl_sheet_range(m):
            full = m.group(0)
            sheet_part, rng = full.split("!", 1)
            sheet = sheet_part[1:-1] if sheet_part.startswith("'") and sheet_part.endswith("'") else sheet_part
            a, b = [x.replace("$", "") for x in rng.split(":")]
            return _ph(f"__range('{sheet}','{a}','{b}')")

        expr = _SHEET_RANGE_RE.sub(repl_sheet_range, expr)

        def repl_range(m):
            rng = m.group(0)
            a, b = [x.replace("$", "") for x in rng.split(":")]
            return _ph(f"__range('{current_sheet}','{a}','{b}')")

        expr = _RANGE_RE.sub(repl_range, expr)

        def repl_sheet_cell(m):
            full = m.group(0)
            sheet_part, addr = full.split("!", 1)
            sheet = sheet_part[1:-1] if sheet_part.startswith("'") and sheet_part.endswith("'") else sheet_part
            return _ph(f"__get('{sheet}','{addr.replace('$','')}')")

        expr = _SHEET_CELL_RE.sub(repl_sheet_cell, expr)

        def repl_cell(m):
            addr = m.group(0).replace("$", "")
            return f"__get('{current_sheet}','{addr}')"

        expr = _CELL_RE.sub(repl_cell, expr)
        for key, repl in placeholders.items():
            expr = expr.replace(key, repl)

        expr = re.sub(r"(?i)\bTRUE\b", "True", expr)
        expr = re.sub(r"(?i)\bFALSE\b", "False", expr)

        def _flatten(args):
            for a in args:
                if isinstance(a, (list, tuple)):
                    for x in a:
                        yield from _flatten([x])
                else:
                    yield a

        def _num(v):
            if isinstance(v, bool):
                return 1.0 if v else 0.0
            n = _to_number(v)
            return 0.0 if n is None else float(n)

        def SUM(*args): return sum(_num(x) for x in _flatten(args))
        def MIN(*args):
            flat = [_num(x) for x in _flatten(args)]
            return min(flat) if flat else 0.0
        def MAX(*args):
            flat = [_num(x) for x in _flatten(args)]
            return max(flat) if flat else 0.0
        def ABS(x): return abs(_num(x))
        def IF(cond, a, b): return a if cond else b
        def IFERROR(a, b):
            try: return a
            except Exception: return b
        def AND(*args): return all(bool(x) for x in _flatten(args))
        def OR(*args): return any(bool(x) for x in _flatten(args))
        def NOT(x): return not bool(x)

        safe_locals = {
            "__get": lambda sheet, addr: self.get_cell(sheet, addr, overrides=overrides),
            "__range": lambda sheet, a, b: self.get_range(sheet, a, b, overrides=overrides),
            "SUM": SUM, "MIN": MIN, "MAX": MAX, "ABS": ABS, "IF": IF, "IFERROR": IFERROR,
            "AND": AND, "OR": OR, "NOT": NOT, "ROUND": round, "POWER": pow,
            "EXP": math.exp, "LN": math.log, "LOG": lambda x, b=10: math.log(x, b), "SQRT": math.sqrt,
        }
        return eval(expr, {"__builtins__": {}}, safe_locals)

def classify_main_rows(workbook_path: str) -> pd.DataFrame:
    model = WorkbookModel.open(workbook_path)
    try:
        ws = model.get_sheet(MAIN_SHEET)
        cols = _find_main_columns(ws)
        rows = []
        last_owner: Optional[str] = None
        for r in range(HEADER_ROW + 1, ws.max_row + 1):
            raw_param = _safe_cell(ws, r, cols["parameter"])
            param_s = "" if raw_param is None else str(raw_param).strip()
            low_raw = _safe_cell(ws, r, cols["low"])
            high_raw = _safe_cell(ws, r, cols["high"])
            proposed_raw = _safe_cell(ws, r, cols["proposed"])
            unit_raw = _safe_cell(ws, r, cols.get("unit"))
            owner_raw = _safe_cell(ws, r, cols.get("owner"))
            desc_raw = _safe_cell(ws, r, cols.get("description"))
            formula_raw = _safe_cell(ws, r, cols.get("formula"))
            formula_text_raw = _safe_cell(ws, r, cols.get("formula_text"))
            owner_defaulted = bool(_is_blank(owner_raw))
            owner_display = _display_owner(last_owner if owner_defaulted else owner_raw)
            if not _is_blank(owner_raw):
                last_owner = str(owner_raw).strip()

            row_type = "skip"
            reason = ""
            if not _parameter_name_looks_real(param_s):
                row_type = "skip"
                reason = "blank_or_non_parameter_name"
            elif _cell_in_wide_merged_band(ws, r, cols["parameter"]):
                row_type = "section_header"
                reason = "wide_merged_band"
            else:
                has_any_bounds_or_value = any(not _is_blank(v) for v in [low_raw, high_raw, proposed_raw])
                has_formula_cell = any(_is_formula(v) for v in [low_raw, high_raw, proposed_raw])
                has_aux = any(not _is_blank(v) for v in [unit_raw, formula_raw, formula_text_raw, owner_raw])
                if not has_any_bounds_or_value and _is_likely_header_label(param_s):
                    row_type = "section_header"
                    reason = "header_like_label"
                elif not has_any_bounds_or_value and not has_aux:
                    row_type = "section_header"
                    reason = "no_values_or_metadata"
                elif has_formula_cell:
                    row_type = "calculated_parameter"
                    reason = "formula_in_value_column"
                else:
                    row_type = "bounded_input"
                    reason = "user_input_like"

            low_num = _to_number(low_raw)
            high_num = _to_number(high_raw)
            proposed_num = _to_number(proposed_raw)

            if row_type == "calculated_parameter":
                # Evaluate formula cells for display if possible.
                if _is_formula(low_raw):
                    try:
                        low_num = _to_number(model.get_cell(MAIN_SHEET, f"C{r}"))
                    except Exception:
                        pass
                if _is_formula(high_raw):
                    try:
                        high_num = _to_number(model.get_cell(MAIN_SHEET, f"D{r}"))
                    except Exception:
                        pass
                if _is_formula(proposed_raw):
                    try:
                        proposed_num = _to_number(model.get_cell(MAIN_SHEET, f"K{r}"))
                    except Exception:
                        pass

            editable_interval = low_num is not None and high_num is not None and low_num <= high_num
            initial_num = proposed_num
            if editable_interval and initial_num is None:
                initial_num = 0.5 * (low_num + high_num)
            candidate_low, candidate_high, candidate_initial, synthetic_bounds = _synthetic_bounds(
                low_num, high_num, initial_num, proposed_num, low_raw, high_raw
            )

            rows.append({
                "Row": r,
                "Parameter": param_s,
                "Low": low_num if low_num is not None else (_to_display(low_raw) if row_type == "bounded_input" else None),
                "High": high_num if high_num is not None else (_to_display(high_raw) if row_type == "bounded_input" else None),
                "Proposed Value": proposed_num if proposed_num is not None else (_to_display(proposed_raw) if row_type == "bounded_input" else None),
                "Low Num": low_num,
                "High Num": high_num,
                "Proposed Value Num": proposed_num,
                "Initial Num": initial_num,
                "Unit": _to_display(unit_raw),
                "Owner": owner_display,
                "Owner Was Defaulted": bool(owner_defaulted and owner_display == DEFAULT_OWNER),
                "Owner UI": _display_owner_ui(owner_display, bool(owner_defaulted and owner_display == DEFAULT_OWNER)),
                "Description": _to_display(desc_raw),
                "Constraint/Formula": _humanize_constraint(low_num, high_num, formula_text_raw, formula_raw),
                "Row Type": row_type,
                "Reason": reason,
                "Editable Interval": bool(editable_interval),
                "Candidate Low": candidate_low,
                "Candidate High": candidate_high,
                "Candidate Initial": candidate_initial,
                "Synthetic Bounds": bool(synthetic_bounds),
            })
        return pd.DataFrame(rows)
    finally:
        model.close()

def load_main_parameters(workbook_path: str) -> pd.DataFrame:
    df = classify_main_rows(workbook_path)
    keep = df["Row Type"].isin(["bounded_input", "calculated_parameter"])
    return df.loc[keep].copy().reset_index(drop=True)

def build_dependency_graph_from_main(workbook_path: str, params_df: Optional[pd.DataFrame] = None) -> nx.DiGraph:
    """Build a directed dependency graph of parameters.

    An edge ``src -> param`` means ``param``'s value is computed from a formula
    that references ``src``. Each edge stores the connecting formula(s) under the
    ``formulas`` attribute so callers can report how two nodes are related.
    """
    if params_df is None:
        params_df = load_main_parameters(workbook_path)
    row_to_param = {int(r["Row"]): str(r["Parameter"]) for _, r in params_df.iterrows()}
    G = nx.DiGraph()
    for _, r in params_df.iterrows():
        G.add_node(str(r["Parameter"]), owner=str(r["Owner"]), row=int(r["Row"]), row_type=str(r["Row Type"]))
    model = WorkbookModel.open(workbook_path)
    try:
        ws = model.get_sheet(MAIN_SHEET)
        cols = _find_main_columns(ws)
        for row_num, param in row_to_param.items():
            formulas = []
            for col_letter, idx in (("C", cols["low"]), ("D", cols["high"]), ("K", cols["proposed"]), ("H", cols.get("formula"))):
                if not idx:
                    continue
                v = ws.cell(row_num, idx).value
                if isinstance(v, str) and v.startswith("="):
                    formulas.append(v)
            # Map each referenced source row to the formula(s) that reference it.
            ref_formulas: Dict[int, List[str]] = {}
            for formula in formulas:
                for x in _ROW_REF_RE.findall(formula):
                    ref_formulas.setdefault(int(x), []).append(formula)
            for rr, fmls in ref_formulas.items():
                src = row_to_param.get(rr)
                if src and src != param:
                    G.add_edge(src, param, formulas=fmls)
    finally:
        model.close()
    return G
